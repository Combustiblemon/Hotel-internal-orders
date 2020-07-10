
from PyQt5 import Qt, QtWidgets, QtCore, uic
from PyQt5.QtWidgets import QLabel, QDialog, QMessageBox, QMainWindow, QWidget, QPushButton, QAction, QInputDialog, QLineEdit, QFileDialog
from PyQt5.QtGui import QIcon
from PyQt5.uic import uiparser
from shutil import copyfile
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Side, Border
from openpyxl.styles import Fill, fills, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.cell import Cell
from copy import copy
import os

uiparser.WidgetStack.topIsLayoutWidget = lambda self: False

borderStyle = Border(left=Side(border_style='medium', color='FF000000'),
                     right=Side(border_style='medium', color='FF000000'),
                     top=Side(border_style='medium', color='FF000000'),
                     bottom=Side(border_style='medium', color='FF000000'),
                     diagonal=Side(border_style=None, color='FF000000'),
                     diagonal_direction=0,
                     outline=Side(border_style=None, color='FF000000'),
                     vertical=Side(border_style=None, color='FF000000'),
                     horizontal=Side(border_style=None, color='FF000000')
                     )      

class MainWindow(QtWidgets.QMainWindow):
    def __init__(self, sectionDictionary, productDictionary):
        super(MainWindow, self).__init__()
        # Load the main UI file
        uic.loadUi('./files/UI/mainWindow.ui', self)
        self.sectionDictionary = sectionDictionary
        self.productDictionary = productDictionary
        self.sectionList = []
        self.text = ''
        self.items = []
        self.index = 0
        
        # create the sectionList from the sectionDictionary
        for k in sectionDictionary.items():
            self.sectionList.append(k[0])
        self.ConnectLogicToObjects()
        self.show()

    # Adding pointers to all the objects of the UI
    def ConnectLogicToObjects(self):
        self.orderView = self.findChild(QtWidgets.QTableWidget, 'orderView')
        self.SetTableStyle()
        # self.AddItemToView('15', 'kg', 'Πόλπα ντομάτα')

        self.addItemButton = self.findChild(QtWidgets.QPushButton, 'addItem')
        self.addItemButton.clicked.connect(self.AddItemPressed)

        self.createOrderButton = self.findChild(QtWidgets.QPushButton, 'createOrder')
        self.createOrderButton.clicked.connect(self.CreateOrderPressed)
        
        self.deleteItemButton = self.findChild(QtWidgets.QPushButton, 'deleteItem')
        self.deleteItemButton.clicked.connect(self.DeleteItemPressed)
        
        self.searchButton = self.findChild(QtWidgets.QPushButton, 'searchButton')
        self.searchButton.clicked.connect(self.searchTable) 
        
        self.clearButton = self.findChild(QtWidgets.QPushButton, 'clearButton')
        self.clearButton.clicked.connect(self.clearButtonPressed)
        
        self.dateLabel = self.findChild(QtWidgets.QLabel, 'dateLabel')
        self.fromLabel = self.findChild(QtWidgets.QLabel, 'fromLabel')
        self.toLabel = self.findChild(QtWidgets.QLabel, 'toLabel')
        
        self.fromSelectorBox = self.findChild(QtWidgets.QComboBox, 'fromSelector')
        self.fromSelectorBox.clear()
        self.fromSelectorBox.addItems(self.sectionList)
        
        self.toSelectorBox = self.findChild(QtWidgets.QComboBox, 'toSelector')
        self.toSelectorBox.clear()
        self.toSelectorBox.addItems(self.sectionList)
        
        self.searchInput = self.findChild(QtWidgets.QLineEdit, 'searchInput')
        
        self.dateEdit = self.findChild(QtWidgets.QDateEdit, 'dateEdit')
        self.SetDateStyle()
        
        self.menuBar = self.findChild(QtWidgets.QMenuBar, 'menubar')
        self.menu = self.findChild(QtWidgets.QMenu, 'menu')
        self.adminMenu = self.findChild(QtWidgets.QMenu, 'administrator')
        self.adminTools = self.findChild(QtWidgets.QAction, 'adminTools')
        self.menuExitOption = self.findChild(QtWidgets.QAction, 'exit')
        self.openOrder = self.findChild(QtWidgets.QAction, 'openOrder')
        self.SetUpMenuBar()
        
    def SetUpMenuBar(self):
        # setup menu exit button       
        self.menuExitOption.setShortcut('Ctrl+Q')
        self.menuExitOption.setStatusTip('Έξοδος προγράμματος')
        self.menuExitOption.triggered.connect(self.exitCall)
        
        # setup admin tools option
        self.adminTools.setShortcut('Ctrl+E')
        self.adminTools.setStatusTip('Εργαλεία Διαχειριστή')
        self.adminTools.triggered.connect(self.adminToolsPressed)
        
        # setup open order option
        self.openOrder.setShortcut('Ctrl+R')
        self.openOrder.setStatusTip('Αλλαγή Υπάρχουσας παραγγελίας')
        self.openOrder.triggered.connect(self.openOrderFile)
        
    def SetTableStyle(self):
        header = self.orderView.horizontalHeader()
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
        
    def SetDateStyle(self):
        self.dateEdit.setDateTime(QtCore.QDateTime.currentDateTime())
        self.dateEdit.setMaximumDate(QtCore.QDate(7999, 12, 28))
        
    def clearButtonPressed(self):
        self.orderView.setRowCount(0)

    def AddItemPressed(self):
        # get the To section from the selection
        selectedSectionTo = str(self.toSelectorBox.currentText())
        
        # open the addItemWindow for the user to select the item
        dialog = AddItemWindow(self.productDictionary[self.sectionDictionary[selectedSectionTo]], selectedSectionTo)
        # get the item selected
        item = dialog.getResults()
        dialog.close()
        if(item is None):
            return
        if(len(item) > 0):
            self.AddItemToView(item[0], item[1], item[2])

    def CreateOrderPressed(self):
        # get the filepath to save the order
        filepath = self.saveFileDialog()
        if filepath is None:
            return
        
        # get the table data of the main window
        data = self.getTableData()
        try:
            # load the template
            wb = load_workbook(filename="./files/data/Template.xlsx")
        except Exception as e:
            QMessageBox.information(self, 'Σφάλμα!', f'Το αρχείο "{e.filename}" δεν υπάρχει.\nΕπικοινωνήστε με τον διαχειρηστή του συστήματος σας.\n\n{e}')
        ws = wb.worksheets[0]
        
        # write the From section to the template
        ws['B2'] = self.fromSelectorBox.currentText()
        
        # write the To section to the template
        ws['D2'] = self.toSelectorBox.currentText()
        
        # write the selected date to the template
        ws['E3'] = self.dateEdit.date().toString('dd/MM/yyyy')
        for index, row in enumerate(data):
            # add the data to the template
            ws[f'A{index + 5}'] = row[0]
            ws[f'B{index + 5}'] = row[1]
            
            # style the cells
            ws[f'A{index + 5}']._style = copy(ws['B3']._style)
            ws[f'B{index + 5}']._style = copy(ws['B3']._style)
            ws[f'C{index + 5}']._style = copy(ws['B3']._style)
            ws[f'D{index + 5}']._style = copy(ws['B3']._style)
            ws[f'E{index + 5}']._style = copy(ws['B3']._style)
        
        try: 
            # save the template as an order to the filepath specified   
            wb.save(filepath)
        except Exception as e:
            QMessageBox.information(self, 'Σφάλμα!', f'Το αρχείο "{e.filename}" ειναι ήδη ανοιχτό.\nΚλείστε το αρχείο και ξαναπροσπαθήστε.\n\n{e}')
    
    def adminToolsPressed(self):
        # ask for a password
        password = PasswordInputWindow().getData()
        if password == os.getenv('ADMIN_PASSWORD'):
            window = adminTools(self.sectionList, self.sectionDictionary)
            window.exec_()
        else:
            QMessageBox.information(self, 'Προσοχή', 'Εσφαλμένος κωδικός')  
            
    def DeleteItemPressed(self):
        # get currently selected item
        rowPosition = self.orderView.currentRow()
        # delete the item
        self.orderView.removeRow(rowPosition)
        
    def exitCall(self):
        self.close()  
        
    def openOrderFile(self):
        # get the filepath of the order to be opened
        filepath = self.openFileDialog()
        
        #if tehre is no filepath return
        if filepath is None:
            return
        
        # load the excel from the filepath
        wb = load_workbook(filename=filepath)
        
        # if nothing was loaded return
        if wb is None:
            return
        # select the first sheet
        ws = wb.worksheets[0]
        
        # clear the table in the Main Window
        self.orderView.setRowCount(0)
        
        # add items from the file to view
        for row in ws.iter_rows(min_row=5, values_only=True):
            # if the row is empty ignore it
            if(row[0] is None):
                continue
            self.AddItemToView(amount=row[0], name=row[1])
        # sort the added items
        self.orderView.sortItems(1, order=QtCore.Qt.AscendingOrder)

    def AddItemToView(self, amount='0', unit='', name=''):
        '''Adds an item to the list.
 
        :param amount: The amount of the item
        :type amount: float
        :param unit: The unit of the item
        :type unit: string
        :param name: The name of the item
        :type name: string
        '''
        totalUnits = f'{amount} {unit}'
        
        # get current row
        rowPosition = self.orderView.rowCount()
        
        # add a new row at the end of the table 
        self.orderView.insertRow(rowPosition)
        
        # populate the row
        self.orderView.setItem(rowPosition, 0, QtWidgets.QTableWidgetItem(totalUnits.strip()))
        self.orderView.setItem(rowPosition, 1, QtWidgets.QTableWidgetItem(name.strip()))
        
    def searchTable(self):
        # if the text in the search bar is not the same as self.text update self.text with the new text
        if (self.text != self.searchInput.text().strip().upper()):
            self.text = self.searchInput.text().strip().upper()
            # find all the items containing self.text and add them to self.items
            self.items = self.orderView.findItems(self.text, QtCore.Qt.MatchContains)
            # set self.index to 0
            self.index = 0
        else:
            # if the index is less than the amount of items, increment it by one to get the next item
            # otherwise set it back to 0
            if(self.index < len(self.items) - 1):
                self.index += 1
            else:
                self.index = 0
        if self.items:
            # select the next item found from the search
            self.orderView.setCurrentItem(self.items[self.index])
        else:
            # if nothing was found and the searchbar wasn't empty display message
            if(self.text != ''):
                QMessageBox.information(self, 'Αναζήτηση', f'Δεν βρέθηκε το προϊόν με το ονομα "{self.searchInput.text().strip()}"')
    
    def getTableData(self):
        # create model from orderView
        model = self.orderView.model()
        data = []
        
        # add all the contents of the model to data[]
        for row in range(model.rowCount()):
            data.append([])
            for column in range(model.columnCount()):
                index = model.index(row, column)
                # We suppose data are strings
                data[row].append(str(model.data(index)))
        return data
    
    # opens the save file dialog and returns the path to the file to be saved as selected by the user
    def saveFileDialog(self, filetypes="Excel Files (*.xlsx);;All Files (*)"):
        options = QFileDialog.Options()
        fileName, _ = QFileDialog.getSaveFileName(self, "Αποθήκευση παραγγελίας", "", filetypes, options=options)
        if fileName:
            return fileName
    
    # opens the open file dialog and returns   teh path to the file to be opened as selected by the user    
    def openFileDialog(self, filetypes="Excel Files (*.xlsx);;All Files (*)"):
        options = QFileDialog.Options()
        fileName, _ = QFileDialog.getOpenFileName(self, "Άνοιγμα λίστας", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        if fileName:
            return fileName
    
    # before the main window closes open a message asking the user if they want to exit 
    def closeEvent(self, event):
        close = QMessageBox()
        close.setWindowTitle('Έξοδος')
        close.setText("Θέλετε να κλείσετε το πρόγραμμα;")
        close.setStandardButtons(QMessageBox.Yes | QMessageBox.Cancel)
        close = close.exec()

        # if the user accepts close the window otherwise ignore the close order
        if close == QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()
            

class AddItemWindow(QtWidgets.QDialog):
    def __init__(self, productDictionary, currentSection):
        super(AddItemWindow, self).__init__()
        # Load the main UI file
        uic.loadUi('./files/UI/createOrder.ui', self)
        self.text = ''
        self.items = []
        self.index = 0
        self.ConnectLogicToObjects(productDictionary, currentSection)   
    
    def ConnectLogicToObjects(self, productDictionary, currentSection):
        self.addButton = self.findChild(QtWidgets.QPushButton, 'addItem')
        self.addButton.clicked.connect(self.AddButtonPressed)
        
        self.searchButton = self.findChild(QtWidgets.QPushButton, 'searchButton')
        self.searchButton.clicked.connect(self.searchTable) 
        
        self.manualAddItemButton = self.findChild(QtWidgets.QPushButton, 'manualAddItem')
        self.manualAddItemButton.clicked.connect(self.manualAddItemPressed) 
        
        self.currentSectionLabel = self.findChild(QtWidgets.QLabel, 'currentSectionLabel')
        self.currentSectionLabel.setText(currentSection)
        
        self.searchInput = self.findChild(QtWidgets.QLineEdit, 'searchInput')
        
        self.productList = self.findChild(QtWidgets.QTableWidget, 'productList')
        self.SetTableStyle()
        
        # adds the items from the selected section to view
        for i in range(len(productDictionary)):
            self.AddItemToView(productDictionary[i][2], productDictionary[i][1])
        # sort the item view
        self.productList.sortItems(1, order=QtCore.Qt.AscendingOrder)
            
    def SetTableStyle(self):
        header = self.productList.horizontalHeader()
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
        
    def searchTable(self):
        # if the text in the search bar is not the same as self.text update self.text with the new text
        if (self.text != self.searchInput.text().strip().upper()):
            self.text = self.searchInput.text().strip().upper()
            # find all the items containing self.text and add them to self.items
            self.items = self.productList.findItems(self.text, QtCore.Qt.MatchContains)
            # set self.index to 0
            self.index = 0
        else:
            # if the index is less than the amount of items, increment it by one to get the next item
            # otherwise set it back to 0
            if(self.index < len(self.items) - 1):
                self.index += 1
            else:
                self.index = 0
        if self.items:
            # select the next item found from the search
            self.productList.selectRow(self.items[self.index].row())
        else:
            # if nothing was found and the searchbar wasn't empty display message
            if(self.text != ''):
                QMessageBox.information(self, 'Αναζήτηση', f'Δεν βρέθηκε το προϊόν με το ονομα "{self.searchInput.text().strip()}"')        
                
    def manualAddItemPressed(self):
        # clear memory before continuing
        self.currentItem = []
        
        # open an input box and get the data input by the user
        inputBox = addExtraItem()
        amount = inputBox.getResults()
        inputBox.close()
        
        # if nothing was input close the window, else add the data to self.currentItems
        if (amount is None):
            self.close()
        else:
            self.currentItem.append(amount[0]) 
            self.currentItem.append(amount[1])
            self.currentItem.append(amount[2])
            
    def AddButtonPressed(self):
        # clear memory before continuing
        self.currentItem = []
        
        # if there is an item selected proceed
        if self.productList.currentItem():
            # open an ItemNumberInput window and get the data input by the user
            inputBox = ItemNumberInput()
            amount = inputBox.getResults()
            inputBox.close()
            
            #if nothing was inputed close the window else add the data to self.currentItem
            if (amount is None):
                self.close()
            else:
                self.currentItem.append(amount) 
                self.currentItem.append(self.productList.item(self.productList.currentItem().row(), 0).text())
                self.currentItem.append(self.productList.item(self.productList.currentItem().row(), 1).text())
                
    def getResults(self):
        # return self.currentItem
        if self.exec_() == QDialog.Accepted:
            return self.currentItem
        else:
            return None
    
    def AddItemToView(self, unit='', name=''):
        '''Adds an item to the list.
 
        :param amount: The amount of the item
        :type amount: float
        :param unit: The unit of the item
        :type unit: string
        :param name: The name of the item
        :type name: string
        '''
        # get current row
        rowPosition = self.productList.rowCount()
        
        # add a new row at the end of the table
        self.productList.insertRow(rowPosition)
        
        # populate the row
        self.productList.setItem(rowPosition, 0, QtWidgets.QTableWidgetItem(f'{unit}'))
        self.productList.setItem(rowPosition, 1, QtWidgets.QTableWidgetItem(name))

class ItemNumberInput(QtWidgets.QDialog):
    def __init__(self):
        super(ItemNumberInput, self).__init__()
        # Load the main UI file
        uic.loadUi('./files/UI/itemNumberInput.ui', self)
        self.ConnectLogicToObjects()
        
    def ConnectLogicToObjects(self):
        self.itemLabel = self.findChild(QtWidgets.QLabel, 'itemLabel')
        self.itemInput = self.findChild(QtWidgets.QLineEdit, 'itemInput')
        self.buttonBox = self.findChild(QtWidgets.QDialogButtonBox, 'buttonBox')
        
    def getResults(self):
        if self.exec_() == QDialog.Accepted:
            return self.itemInput.text()
        else:
            return None
    
class addExtraItem(QtWidgets.QDialog):
    def __init__(self):
        super(addExtraItem, self).__init__()
        # Load the main UI file
        uic.loadUi('./files/UI/addExtraItem.ui', self)
        self.ConnectLogicToObjects()
        
    def ConnectLogicToObjects(self):
        self.nameInputBox = self.findChild(QtWidgets.QLineEdit, 'nameInputBox')
        self.amountInputBox = self.findChild(QtWidgets.QLineEdit, 'amountInputBox')
        self.unitInputBox = self.findChild(QtWidgets.QLineEdit, 'unitInputBox')
        
        self.amountLabel = self.findChild(QtWidgets.QLabel, 'amountLabel')
        self.nameLabel = self.findChild(QtWidgets.QLabel, 'nameLabel')
        self.unitInputLabel = self.findChild(QtWidgets.QLabel, 'unitInputLabel')
        
    def getResults(self):
        if self.exec_() == QDialog.Accepted:
            item = ['0', '0', 'Empty']
            item[0] = self.amountInputBox.text().strip().upper()
            item[1] = self.unitInputBox.text().strip().upper()
            item[2] = self.nameInputBox.text().strip().upper()
            return item
        else:
            return None
        
class adminTools(QtWidgets.QDialog):
    def __init__(self, sectionList, sectionDictionary):
        super(adminTools, self).__init__()
        # Load the main UI file
        uic.loadUi('./files/UI/adminTools.ui', self)
        self.sectionList = sectionList
        self.sectionDictionary = sectionDictionary
        self.ConnectLogicToObjects()
        
    def ConnectLogicToObjects(self):
        self.changeProductList = self.findChild(QtWidgets.QPushButton, 'changeProductList')
        self.changeProductList.clicked.connect(self.changeProductListClicked)
    
        self.addSingleItem = self.findChild(QtWidgets.QPushButton, 'addSingleItem')
        self.addSingleItem.clicked.connect(self.addSingleItemClicked)
        
        self.adminPasswordChange = self.findChild(QtWidgets.QPushButton, 'adminPasswordChange')
        self.adminPasswordChange.clicked.connect(self.adminPasswordChangeClicked)
        
    def changeProductListClicked(self):
        # get the filepath to the new productList as input by the user
        filepath = self.openFileDialog()
        if(filepath):
            dest = './files/data/ProductList.xlsx'
            try:
                # replace the product list with the new file    
                copyfile(filepath, dest)
            except Exception as e:
                QMessageBox.information(self, 'Σφάλμα!', f'Το αρχείο "{e.filename}" ειναι ήδη ανοιχτό.\nΚλείστε το αρχείο και ξαναπροσπαθήστε.\n\n{e}')
                
    def addSingleItemClicked(self):
        # open an AddExtraItemToData window and get the data input by the user
        dialog = AddExtraItemToData(self.sectionList, self.sectionDictionary)
        item = dialog.getData()
        if item:
            try:
                # load the ProductList
                wb = load_workbook(filename="./files/data/ProductList.xlsx")
            except Exception as e:
                QMessageBox.information(self, 'Σφάλμα!', f'Το αρχείο "{e.filename}" ειναι ήδη ανοιχτό.\nΕπικοινωνήστε με τον διαχειρηστή του συστήματος σας.\n\n{e}')
            
            # add item to total view
            ws = wb.worksheets[0]
            ws.append([item[0], item[1], item[2]])
            
            # add the item to the section worksheet
            if (self.sectionDictionary[item[3]] != wb.sheetnames[0]):
                ws = wb[self.sectionDictionary[item[3]]]
                ws.append([item[0], item[1], item[2]])
            
            try:
                # save ProductList
                wb.save('./files/data/ProductList.xlsx')
            except Exception as e:
                QMessageBox.information(self, 'Σφάλμα!', f'Το αρχείο "{e.filename}" δεν μπορούσε να αποθηκευτεί.\nΕπικοινωνήστε με τον διαχειρηστή του συστήματος σας.\n\n{e}')               
    
    def adminPasswordChangeClicked(self):
        # open a PasswordInputWindow and get the new password as input by the user
        dialog = PasswordInputWindow('Εισαγωγή νέου κωδικού Διαχειριστή:')
        password = dialog.getData()
        if password:
            os.environ["ADMIN_PASSWORD"] = password
            try:
                # write the password to data.dat
                with open('./files/data/data.dat', 'w') as file:
                    file.write(password)
                QMessageBox.information(self, '', 'Ο κωδικός αλλαξε επιτυχώς.')
            except Exception as e:
                QMessageBox.information(self, 'Σφάλμα!', f'Ο κωδικός δεν μπόρεσε να αλλαχθεί. Προσπαθήστε ξανά.\n\n {e}')
        elif password == 'PASSWORD REJECTED':
            return
        else:
            QMessageBox.information(self, 'Σφάλμα!', 'Ο κωδικός δεν μπόρεσε να αλλαχθεί. Προσπαθήστε ξανά.')
    
    # opens the open file dialog and returns   teh path to the file to be opened as selected by the user   
    def openFileDialog(self, filetypes="Excel Files (*.xlsx);;All Files (*)"):
        options = QFileDialog.Options()
        fileName, _ = QFileDialog.getOpenFileName(self, "Άνοιγμα λίστας", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        if fileName:
            return fileName
            
        
class AddExtraItemToData(QtWidgets.QDialog):
    def __init__(self, sectionList, sectionDictionary):
        super(AddExtraItemToData, self).__init__()
        # Load the main UI file
        uic.loadUi('./files/UI/addExtraItemToData.ui', self)
        self.sectionDictionary = sectionDictionary
        self.sectionList = sectionList
        self.ConnectLogicToObjects()
        
    def ConnectLogicToObjects(self):
        self.nameInputBox = self.findChild(QtWidgets.QLineEdit, 'nameInputBox')
        self.codeInputBox = self.findChild(QtWidgets.QLineEdit, 'codeInputBox')
        self.unitInputBox = self.findChild(QtWidgets.QLineEdit, 'unitInputBox')
        
        self.codeLabel = self.findChild(QtWidgets.QLabel, 'codeLabel')
        self.nameLabel = self.findChild(QtWidgets.QLabel, 'nameLabel')
        self.unitInputLabel = self.findChild(QtWidgets.QLabel, 'unitInputLabel')
        self.sectionLabel = self.findChild(QtWidgets.QLabel, 'sectionLabel')
        
        self.sectionSelector = self.findChild(QtWidgets.QComboBox, 'sectionSelector')
        self.sectionSelector.clear()
        self.sectionSelector.addItems(self.sectionList)
        
    def getData(self):
        '''Returns the new item in an array
        
        [0] the code of the item as input by the user
        
        [1] the name of the item as input by the user
        
        [2] the unit of the item as input by the user
        
        [3] the section of the item as input by the user
        '''
        if self.exec_() == QDialog.Accepted:
            item = []
            item.append(self.codeInputBox.text().strip())
            item.append(self.nameInputBox.text().strip())
            item.append(self.unitInputBox.text().strip())
            item.append(self.sectionSelector.currentText())
            return(item)
        else:
            return None
        
class PasswordInputWindow(QtWidgets.QDialog):
    def __init__(self, text='Κωδικός Διαχειριστή:'):
        super(PasswordInputWindow, self).__init__()
        # Load the main UI file
        uic.loadUi('./files/UI/passwordInput.ui', self)
        self.ConnectLogicToObjects(text)
    
    def ConnectLogicToObjects(self, text):
        self.label = self.findChild(QtWidgets.QLabel, 'label')
        self.label.setText(text)
        self.passwordInput = self.findChild(QtWidgets.QLineEdit, 'passwordInput')
        
    def getData(self):
        if self.exec_() == QDialog.Accepted:
            return self.passwordInput.text()
        elif self.exec_() == QDialog.Rejected:
            return 'PASSWORD REJECTED'
        else:
            return None
    
    def closeEvent(self, event):
        self.close()
    
        
class LoadingWindow(QtWidgets.QDialog):
    def __init__(self):
        super(LoadingWindow, self).__init__()
        # Load the main UI file
        uic.loadUi('./files/UI/LoadingWindow.ui', self)