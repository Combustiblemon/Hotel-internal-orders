"""Doc."""
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QMessageBox
from UI import LoadingWindow, MainWindow
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
from pathlib import Path
import os
import sys
import json
import io

# Loading window
app = QtWidgets.QApplication(sys.argv)
loadingW = LoadingWindow()
loadingW.show()
correctLoad = True

# Open data.dat file and get the password
try:
    with open('./files/data/data.dat', 'r') as file:
        os.environ['ADMIN_PASSWORD'] = file.read()
except Exception as e:
    with open('./files/data/data.dat', 'w') as file:
        file.write('1234')

# load the excel file containing all the items
try:
    workbook = load_workbook(filename="./files/data/ProductList.xlsx")
except Exception as e:
    QMessageBox.information(loadingW, 'Σφάλμα!', f'Το αρχείο {e.filename} δεν υπάρχει.\nΕπικοινωνήστε με τον διαχειρηστή του συστήματος σας.\n\n{e}')
    correctLoad = False

# add all the individual sheets into one dictionary and link their data
productList = {}
for sheet in workbook.sheetnames:
    productList[sheet] = []
    for row in workbook[sheet].iter_rows(min_row=3, values_only=True):
        if(row[0] is None):
            continue
        productList[sheet].append(row)

# import the dictionary from the .json file
sectionDictionary = {}
try:
    with open('./files/data/dictionary.json', 'r', encoding="utf8") as file:
        sectionDictionary = json.load(file)
except Exception as e:
    QMessageBox.information(loadingW, 'Σφάλμα!', f'Το αρχείο {e.filename} δεν υπάρχει.\nΕπικοινωνήστε με τον διαχειρηστή του συστήματος σας.\n\n{e}')
    correctLoad = False

# setup the application
if correctLoad:
    loadingW.close()
    window = MainWindow(sectionDictionary, productList)
    window.show()
    app.exec_()